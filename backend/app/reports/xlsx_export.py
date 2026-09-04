import io
import os
from datetime import datetime

from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from PIL import Image as PILImage
from sqlalchemy.orm import Session

from .. import models
from ..config import settings
from .common import CATEGORY_LABELS, JUSTIFICATION_TYPE_LABELS

THUMB_SIZE = (110, 110)
THUMB_ROW_HEIGHT = 85
THUMB_COL_WIDTH = 18

HEADER_FILL = PatternFill("solid", fgColor="123B40")
HEADER_FONT = Font(bold=True, color="FFFFFF")
SUBTOTAL_FILL = PatternFill("solid", fgColor="E4EEEC")
TOTAL_FILL = PatternFill("solid", fgColor="C7D8D5")
BOLD = Font(bold=True)

PAYROLL_HEADER = [
    "Empleado", "Categoría", "Fecha", "Día",
    "Entrada", "Salida a comer", "Regreso de comer", "Salida",
    "Horas trabajadas", "Retardo (min)", "Comida tarde (min)", "Sin marca", "Justificante",
]

PHOTOS_HEADER = ["Empleado", "Fecha", "Día", "Tipo de marca", "Hora", "Archivo de foto"]

DOW_NAMES = ["Lunes", "Martes", "Miércoles", "Jueves", "Viernes", "Sábado", "Domingo"]


def _merge_absences(rows, absences):
    merged = list(rows)
    for a in absences:
        merged.append({
            "employeeId": a["employeeId"], "employeeName": a["employeeName"], "date": a["date"],
            "entrada": None, "comida_salida": None, "comida_entrada": None, "salida": None,
            "hours": 0.0, "retardoMin": 0, "lunchLateMin": 0, "missing": True,
            "justified": a["justified"],
            "justificationType": (
                JUSTIFICATION_TYPE_LABELS.get(a["justificationType"], a["justificationType"])
                if a["justificationType"] else None
            ),
        })
    return merged


def _style_row(ws, row_idx, font=None, fill=None):
    for cell in ws[row_idx]:
        if font:
            cell.font = font
        if fill:
            cell.fill = fill


def _write_payroll_sheet(ws, db, rows, absences, period, date_str, emp_filter):
    period_label = {"dia": "Día", "semana": "Semana", "quincena": "Quincena"}.get(period, period)
    emp_label = "Todos" if (not emp_filter or emp_filter == "all") else emp_filter
    anchor = date_str or datetime.now().strftime("%Y-%m-%d")
    categories = {e.id: e.category for e in db.query(models.Employee).all()}

    ws.append(["Reporte de Asistencia — Reloj Checador"])
    ws["A1"].font = Font(bold=True, size=13)
    ws.append([f"Periodo: {period_label}"])
    ws.append([f'Fecha de referencia: {datetime.strptime(anchor, "%Y-%m-%d").strftime("%d/%m/%Y")}'])
    ws.append([f"Empleado: {emp_label}"])
    ws.append([f'Generado: {datetime.now().strftime("%d/%m/%Y %H:%M")}'])
    ws.append([])

    ws.append(PAYROLL_HEADER)
    _style_row(ws, ws.max_row, font=HEADER_FONT, fill=HEADER_FILL)

    all_rows = _merge_absences(rows, absences)
    sorted_rows = sorted(all_rows, key=lambda r: (r["employeeName"], r["date"]))

    current_emp = None
    emp_hours = 0.0
    emp_missing = emp_unjustified = 0

    def flush_subtotal():
        if current_emp is not None:
            just_summary = f"{emp_unjustified} injustificada(s)" if emp_unjustified else ""
            ws.append([
                f"Subtotal {current_emp}", "", "", "", "", "", "", "",
                round(emp_hours, 2), "", "",
                emp_missing if emp_missing else "",
                just_summary,
            ])
            _style_row(ws, ws.max_row, font=BOLD, fill=SUBTOTAL_FILL)
            ws.append([])

    for r in sorted_rows:
        if current_emp is not None and r["employeeName"] != current_emp:
            flush_subtotal()
            emp_hours = 0.0
            emp_missing = emp_unjustified = 0
        current_emp = r["employeeName"]
        emp_hours += r["hours"]
        if r["missing"]:
            emp_missing += 1
        if r.get("justified") is False:
            emp_unjustified += 1

        fecha_dt = datetime.strptime(r["date"], "%Y-%m-%d")
        cat = CATEGORY_LABELS.get(categories.get(r["employeeId"]), "")

        justificante = ""
        if r.get("justified") is True:
            justificante = f"Sí — {r.get('justificationType') or ''}".rstrip(" —")
        elif r.get("justified") is False:
            justificante = "No"

        ws.append([
            r["employeeName"], cat, fecha_dt.strftime("%d/%m/%Y"), DOW_NAMES[fecha_dt.weekday()],
            r["entrada"] or "—", r["comida_salida"] or "—", r["comida_entrada"] or "—", r["salida"] or "—",
            r["hours"],
            r["retardoMin"] or "", r["lunchLateMin"] or "",
            "Sí" if r["missing"] else "",
            justificante,
        ])
    flush_subtotal()

    total_hours = sum(r["hours"] for r in all_rows)
    total_missing = sum(1 for r in all_rows if r["missing"])
    total_unjustified = sum(1 for r in all_rows if r.get("justified") is False)
    ws.append([
        "TOTAL GENERAL", "", "", "", "", "", "", "",
        round(total_hours, 2), "", "",
        total_missing if total_missing else "",
        f"{total_unjustified} injustificada(s)" if total_unjustified else "",
    ])
    _style_row(ws, ws.max_row, font=BOLD, fill=TOTAL_FILL)

    for i, w in enumerate([22, 14, 12, 11, 9, 13, 14, 9, 11, 11, 13, 10, 24], start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def _thumbnail_image(path):
    try:
        img = PILImage.open(path)
        img.thumbnail(THUMB_SIZE)
        buf = io.BytesIO()
        img.convert("RGB").save(buf, format="JPEG")
        buf.seek(0)
        return XLImage(buf)
    except Exception:
        return None


def _write_photos_sheet(ws, photo_log):
    header = PHOTOS_HEADER + ["Foto"]
    ws.append(header)
    _style_row(ws, 1, font=HEADER_FONT, fill=HEADER_FILL)
    photo_col = get_column_letter(len(header))

    for p in photo_log:
        fecha_dt = datetime.strptime(p["date"], "%Y-%m-%d")
        ws.append([
            p["employeeName"], fecha_dt.strftime("%d/%m/%Y"), p["dayLabel"],
            p["type"], p["time"] or "—", p["photoFile"] or "", "",
        ])
        row_idx = ws.max_row
        ws.row_dimensions[row_idx].height = THUMB_ROW_HEIGHT

        if p.get("photoFile"):
            path = os.path.join(settings.punch_photos_dir, p["photoFile"])
            if os.path.isfile(path):
                thumb = _thumbnail_image(path)
                if thumb:
                    ws.add_image(thumb, f"{photo_col}{row_idx}")

    for i, w in enumerate([22, 12, 11, 16, 8, 34], start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.column_dimensions[photo_col].width = THUMB_COL_WIDTH


def build_xlsx(db: Session, rows, absences, photo_log, period, date_str, emp_filter):
    wb = Workbook()
    ws_payroll = wb.active
    ws_payroll.title = "Nómina"
    _write_payroll_sheet(ws_payroll, db, rows, absences, period, date_str, emp_filter)

    ws_photos = wb.create_sheet("Fotos")
    _write_photos_sheet(ws_photos, photo_log)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
